"""Excel workbook parsing for project health reporting."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

from .utils import normalize

LOGGER = logging.getLogger("project_health_agent")


def load_workbook_data(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    """Load summary, task, and comment data from a project workbook."""
    LOGGER.info("Loading workbook: %s", path)
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    summary = read_summary(workbook)
    comments = read_comments(workbook)
    project_sheet_name = choose_project_sheet(workbook)
    tasks, warnings = read_tasks(workbook[project_sheet_name])
    warnings.extend(find_missing_core_fields(tasks))
    return summary, tasks, comments, warnings


def choose_project_sheet(workbook: openpyxl.Workbook) -> str:
    """Choose the primary task-plan sheet from a workbook."""
    for name in workbook.sheetnames:
        if name.lower() not in {"summary", "comments"}:
            return name
    return workbook.sheetnames[0]


def read_summary(workbook: openpyxl.Workbook) -> dict[str, Any]:
    """Read key-value project summary fields from the Summary sheet."""
    if "Summary" not in workbook.sheetnames:
        return {}
    summary: dict[str, Any] = {}
    for row in workbook["Summary"].iter_rows(values_only=True):
        key = normalize(row[0] if row else None)
        if key:
            summary[key] = row[1] if len(row) > 1 else None
    return summary


def read_comments(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    """Read PM comments and metadata from the Comments sheet."""
    if "Comments" not in workbook.sheetnames:
        return []
    rows: list[dict[str, Any]] = []
    for row in workbook["Comments"].iter_rows(values_only=True):
        if not row or not any(cell is not None for cell in row):
            continue
        rows.append(
            {
                "row_ref": normalize(row[0] if len(row) > 0 else ""),
                "comment": normalize(row[1] if len(row) > 1 else ""),
                "author": normalize(row[2] if len(row) > 2 else ""),
                "timestamp": normalize(row[3] if len(row) > 3 else ""),
            }
        )
    return rows


def read_tasks(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[dict[str, Any]], list[str]]:
    """Read task rows from the selected project-plan worksheet."""
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [normalize(cell) for cell in next(rows)]
    except StopIteration:
        return [], ["Project sheet is empty."]

    tasks: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        record = {"_row": row_number}
        for index, header in enumerate(headers):
            if header:
                record[header] = row[index] if index < len(row) else None
        if normalize(record.get("Task Name")) or normalize(record.get("Status")):
            tasks.append(record)

    warnings = [] if tasks else [f"No task rows found in sheet '{sheet.title}'."]
    return tasks, warnings


def find_missing_core_fields(tasks: list[dict[str, Any]]) -> list[str]:
    """Return data-quality warnings for missing core task columns."""
    available = set().union(*(task.keys() for task in tasks)) if tasks else set()
    expected = ("Task Name", "Status", "Schedule Health", "Start Date", "End Date", "% Complete")
    return [f"Missing expected column: {column}." for column in expected if column not in available]

